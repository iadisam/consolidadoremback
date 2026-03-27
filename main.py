"""
API REST para Sistema de Consolidación REM
FastAPI + SQL Server + pyodbc
Versión: 2.5.0
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, status, Form
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime, timedelta
import pyodbc
import bcrypt
import jwt
import os
from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import re

# Cargar variables de entorno
from dotenv import load_dotenv
load_dotenv()

# ==================== CONFIGURACIÓN ====================
app = FastAPI(
    title="API Consolidador REM",
    description="API para gestión de archivos REM y consolidación",
    version="2.5.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración JWT
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise ValueError("❌ ERROR: SECRET_KEY no configurada en .env")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "480"))

# Configuración SQL Server
DB_CONFIG = {
    "DRIVER": os.getenv("DB_DRIVER", "{ODBC Driver 18 for SQL Server}"),
    "SERVER": os.getenv("DB_SERVER"),
    "DATABASE": os.getenv("DB_NAME"),
    "UID": os.getenv("DB_USER"),
    "PWD": os.getenv("DB_PASSWORD"),
    "TrustServerCertificate": "yes", 
    "Encrypt": "yes",
}

# Validar variables requeridas
required_vars = ["DB_SERVER", "DB_NAME", "DB_USER", "DB_PASSWORD"]
missing_vars = [var for var in required_vars if not os.getenv(var)]
if missing_vars:
    raise ValueError(f"❌ ERROR: Variables faltantes en .env: {', '.join(missing_vars)}")

# Mostrar configuración (sin password)
print("="*60)
print(" CONFIGURACIÓN DE BASE DE DATOS:")
print(f"   SERVER: {DB_CONFIG['SERVER']}")
print(f"   DATABASE: {DB_CONFIG['DATABASE']}")
print(f"   USER: {DB_CONFIG['UID']}")
print(f"   PASSWORD: {'*' * len(DB_CONFIG['PWD']) if DB_CONFIG['PWD'] else 'NO CONFIGURADA'}")
print("="*60)

# Directorio de archivos
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Directorio temporal
TEMP_UPLOAD_DIR = Path("temp_uploads")
TEMP_UPLOAD_DIR.mkdir(exist_ok=True)

# Security
security = HTTPBearer()

# ====================MODELOS PYDANTIC ====================

class ProgramaBase(BaseModel):
    nombre: str

class ProgramaCreate(ProgramaBase):
    pass

class Programa(ProgramaBase):
    id: int
    activo: bool
    
    class Config:
        from_attributes = True

class UsuarioCreate(BaseModel):
    nombre: str
    email: EmailStr
    password: str
    rol: str
    programa_id: Optional[int] = None

class UsuarioLogin(BaseModel):
    email: EmailStr
    password: str

class Usuario(BaseModel):
    id: int
    nombre: str
    email: str
    rol: str
    programa_id: Optional[int]
    activo: bool
    created_at: datetime
    programa_nombre: Optional[str]
    
    class Config:
        from_attributes = True

class ArchivoValidar(BaseModel):
    archivo_id: int
    estado: str
    observaciones: Optional[str] = None

class ConsolidarRequest(BaseModel):
    archivos_ids: List[int]
    periodo: str

class ConsolidacionInfo(BaseModel):
    id: int
    nombre_archivo: str
    archivos_count: int
    creado_por: int
    fecha: datetime
    periodo: str
    creado_por_nombre: str
    
    class Config:
        from_attributes = True

# ==================== UTILIDADES ====================

def get_db_connection():
    """Crea conexión a SQL Server"""
    try:
        conn_str = ";".join([f"{k}={v}" for k, v in DB_CONFIG.items()])
        conn = pyodbc.connect(conn_str)
        return conn
    except Exception as e:
        print(f"❌ Error conectando a base de datos: {e}")
        raise HTTPException(status_code=500, detail=f"Error de conexión a base de datos: {str(e)}")

def hash_password(password: str) -> str:
    """Hashea password con bcrypt"""
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verifica password con bcrypt"""
    return bcrypt.checkpw(plain_password.encode(), hashed_password.encode())

def create_access_token(data: dict):
    """Crea JWT token"""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def decode_token(token: str):
    """Decodifica JWT token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Obtener usuario actual desde token"""
    token = credentials.credentials
    payload = decode_token(token)
    
    user_id = payload.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.rol, u.programa_id, u.activo, 
               p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        WHERE u.id = ?
    """, user_id)
    
    row = cursor.fetchone()
    conn.close()
    
    if not row or not row[5]:
        raise HTTPException(status_code=401, detail="Usuario no encontrado o inactivo")
    
    return {
        "id": row[0],
        "nombre": row[1],
        "email": row[2],
        "rol": row[3],
        "programa_id": row[4],
        "programa_nombre": row[6]
    }

def registrar_log(usuario_id: int, accion: str, detalle: str = None, 
                  archivo_id: int = None, consolidacion_id: int = None):
    """Registra actividad en log"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO log_actividad (usuario_id, accion, detalle, archivo_id, consolidacion_id)
            VALUES (?, ?, ?, ?, ?)
        """, usuario_id, accion, detalle, archivo_id, consolidacion_id)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️  Error registrando log: {e}")

def get_periodo_actual() -> str:
    """Retorna el período actual en formato YYYY-MM"""
    now = datetime.now()
    return now.strftime("%Y-%m")

def get_upload_dir_for_periodo(periodo: str) -> Path:
    """Retorna el directorio de uploads para un período específico"""
    periodo_dir = UPLOAD_DIR / periodo
    periodo_dir.mkdir(exist_ok=True)
    return periodo_dir

def validar_archivo_rem(filepath: Path, periodo_esperado: str) -> dict:
    """
    Valida archivo .xlsm:
    - 30 hojas
    - Versión 1.2
    - Mes del archivo corresponde al periodo esperado
    """
    MESES_MAP = {
        "01": "ENERO", "02": "FEBRERO", "03": "MARZO", "04": "ABRIL",
        "05": "MAYO", "06": "JUNIO", "07": "JULIO", "08": "AGOSTO",
        "09": "SEPTIEMBRE", "10": "OCTUBRE", "11": "NOVIEMBRE", "12": "DICIEMBRE"
    }
    
    try:
        wb = load_workbook(filepath, data_only=True, keep_vba=True)
        
        # Validar 30 hojas
        if len(wb.sheetnames) != 30:
            wb.close()
            return {
                "valido": False,
                "error": "Validación de estructura fallida",
                "mensaje": f"El archivo tiene {len(wb.sheetnames)} hojas, se esperaban 30"
            }
        
        # Validar hoja NOMBRE y versión
        if 'NOMBRE' not in wb.sheetnames:
            wb.close()
            return {
                "valido": False,
                "error": "Validación de estructura fallida",
                "mensaje": "No se encontró la hoja NOMBRE"
            }
        
        ws_nombre = wb['NOMBRE']
        
        # Validar versión (buscar "Versión 1.2" o "VERSION 1.2" en la hoja)
        version_encontrada = False
        for row in ws_nombre.iter_rows(max_row=15):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    valor_upper = cell.value.upper()
                    if 'VERSIÓN 1.2' in valor_upper or 'VERSION 1.2' in valor_upper:
                        version_encontrada = True
                        break
            if version_encontrada:
                break
        
        if not version_encontrada:
            wb.close()
            return {
                "valido": False,
                "error": "Validación de versión fallida",
                "mensaje": "El archivo no corresponde a la versión 1.2"
            }
        
        # Validar mes del archivo - Extraer de celda A9 "Versión 1.2: Febrero 2026"
        celda_a9 = ws_nombre['A9'].value
        
        if celda_a9 and isinstance(celda_a9, str):
            # Extraer mes y año de texto como "Versión 1.2: Febrero 2026"
            match = re.search(r'(\w+)\s+(\d{4})', celda_a9)
            if match:
                mes_archivo_valor = match.group(1).upper()
                anio_archivo_valor = match.group(2)
            else:
                mes_archivo_valor = None
                anio_archivo_valor = None
        else:
            mes_archivo_valor = None
            anio_archivo_valor = None
        
        wb.close()
        
        if not mes_archivo_valor or not anio_archivo_valor:
            return {
                "valido": False,
                "error": "Validación de mes fallida",
                "mensaje": "No se encontró información de mes/año en el archivo (celda A9)"
            }
        
        # Parsear periodo esperado
        anio_esperado, mes_esperado_num = periodo_esperado.split('-')
        mes_esperado_nombre = MESES_MAP[mes_esperado_num]
        
        # El mes ya está en mayúsculas y normalizado
        mes_archivo_str = mes_archivo_valor
        
        try:
            anio_archivo_int = int(anio_archivo_valor)
        except:
            return {
                "valido": False,
                "error": "Validación de mes fallida",
                "mensaje": f"El año en el archivo ('{anio_archivo_valor}') no es válido"
            }
        
        # Validar coincidencia
        if mes_archivo_str != mes_esperado_nombre or anio_archivo_int != int(anio_esperado):
            return {
                "valido": False,
                "error": "Validación de mes fallida",
                "mensaje": f"El archivo corresponde a {mes_archivo_str} {anio_archivo_int} pero se esperaba {mes_esperado_nombre} {anio_esperado}",
                "mes_archivo": mes_archivo_str,
                "anio_archivo": anio_archivo_int,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": int(anio_esperado),
                "ayuda": f"Recuerde: debe subir el reporte del período {periodo_esperado}"
            }
        
        return {
            "valido": True,
            "mensaje": f"Archivo válido: {mes_archivo_str} {anio_archivo_int}"
        }
        
    except Exception as e:
        return {
            "valido": False,
            "error": "Error al validar archivo",
            "mensaje": str(e)
        }


# ==================== ENDPOINTS ====================

@app.get("/")
def root():
    """Health check"""
    return {
        "status": "ok",
        "service": "API Consolidador REM",
        "version": "2.5.0",
        "docs": "/docs"
    }

@app.get("/health")
def health_check():
    """Verificar estado de la API"""
    try:
        conn = get_db_connection()
        conn.close()
        return {"status": "healthy", "database": "connected", "version": "2.5.0"}
    except:
        return {"status": "unhealthy", "database": "disconnected", "version": "2.5.0"}

# ==================== PLANTILLA ====================

@app.get("/plantilla/download")
def descargar_plantilla():
    """Descargar plantilla base SA_26_V1.2.xlsm (sin autenticación)"""
    plantilla_path = Path("SA_26_V1.2.xlsm")
    
    if not plantilla_path.exists():
        raise HTTPException(status_code=404, detail="Plantilla no encontrada. Coloque SA_26_V1.2.xlsm en la raíz del proyecto")
    
    return FileResponse(
        path=str(plantilla_path),
        filename="SA_26_V1.2.xlsm",
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

# ==================== AUTENTICACIÓN ====================

@app.post("/auth/register")
def registrar_usuario(usuario: UsuarioCreate):
    """Registrar nuevo usuario"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar si email ya existe
    cursor.execute("SELECT id FROM usuarios WHERE email = ?", usuario.email)
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Email ya registrado")
    
    # Hashear password
    password_hash = hash_password(usuario.password)
    
    # Insertar usuario
    cursor.execute("""
        INSERT INTO usuarios (nombre, email, password_hash, rol, programa_id)
        VALUES (?, ?, ?, ?, ?)
    """, usuario.nombre, usuario.email, password_hash, usuario.rol, usuario.programa_id)
    
    conn.commit()
    user_id = cursor.execute("SELECT @@IDENTITY").fetchone()[0]
    conn.close()
    
    return {"message": "Usuario creado exitosamente", "user_id": user_id}

@app.post("/auth/login")
def login(credentials: UsuarioLogin):
    """Login y generar token"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.password_hash, u.rol, u.programa_id, u.activo,
               p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        WHERE u.email = ?
    """, credentials.email)
    
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    if not verify_password(credentials.password, row[3]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    if not row[6]:
        raise HTTPException(status_code=401, detail="Usuario inactivo")
    
    token = create_access_token({
        "user_id": row[0],
        "email": row[2],
        "rol": row[4]
    })
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": row[0],
            "nombre": row[1],
            "email": row[2],
            "rol": row[4],
            "programa_id": row[5],
            "programa_nombre": row[7]
        }
    }

@app.get("/auth/me")
def get_me(current_user: dict = Depends(get_current_user)):
    """Obtener info del usuario actual"""
    return current_user

@app.get("/usuarios", response_model=List[Usuario])
def listar_usuarios(current_user: dict = Depends(get_current_user)):
    """Listar usuarios encargados con nombre de programa (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.rol, u.programa_id, u.activo, u.created_at,
               p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        WHERE u.rol = 'encargado'
        ORDER BY u.nombre
    """)
    
    usuarios = []
    for row in cursor.fetchall():
        usuarios.append({
            "id": row[0],
            "nombre": row[1],
            "email": row[2],
            "rol": row[3],
            "programa_id": row[4],
            "activo": row[5],
            "created_at": row[6],
            "programa_nombre": row[7]
        })
    
    conn.close()
    return usuarios

# ==================== PROGRAMAS ====================

@app.get("/programas")
def listar_programas(current_user: dict = Depends(get_current_user)):
    """Listar programas activos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, nombre, codigo FROM programas WHERE activo = 1 ORDER BY nombre")
    
    programas = []
    for row in cursor.fetchall():
        programas.append({
            "id": row[0],
            "nombre": row[1],
            "codigo": row[2]
        })
    
    conn.close()
    return programas

# ==================== ARCHIVOS ====================

@app.get("/archivos")
def listar_archivos(
    estado: Optional[str] = None,
    programa_id: Optional[int] = None,
    periodo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar archivos con filtros opcionales"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT 
            a.id, a.usuario_id, a.programa_id, a.nombre_archivo, a.estado,
            a.observaciones, a.fecha_subida, a.fecha_validacion, a.validado_por,
            a.activo, a.periodo, u.nombre as usuario_nombre, p.nombre as programa_nombre,
            v.nombre as validado_por_nombre
        FROM archivos a
        JOIN usuarios u ON a.usuario_id = u.id
        JOIN programas p ON a.programa_id = p.id
        LEFT JOIN usuarios v ON a.validado_por = v.id
        WHERE a.activo = 1
    """
    params = []
    
    if estado:
        query += " AND a.estado = ?"
        params.append(estado)
    
    if programa_id:
        query += " AND a.programa_id = ?"
        params.append(programa_id)
    
    if periodo:
        query += " AND a.periodo = ?"
        params.append(periodo)
    
    if current_user["rol"] == "encargado":
        query += " AND a.usuario_id = ?"
        params.append(current_user["id"])
    
    query += " ORDER BY a.fecha_subida DESC"
    
    if params:
        cursor.execute(query, *params)
    else:
        cursor.execute(query)
    
    archivos = []
    for row in cursor.fetchall():
        archivos.append({
            "id": row[0],
            "usuario_id": row[1],
            "programa_id": row[2],
            "nombre_archivo": row[3],
            "estado": row[4],
            "observaciones": row[5],
            "fecha_subida": row[6],
            "fecha_validacion": row[7],
            "validado_por": row[8],
            "activo": row[9],
            "periodo": row[10],
            "usuario_nombre": row[11],
            "programa_nombre": row[12],
            "validado_por_nombre": row[13]
        })
    
    conn.close()
    return archivos

@app.post("/archivos/upload")
async def subir_archivo(
    file: UploadFile = File(...),
    programa_id: Optional[int] = Form(None),
    periodo: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Subir archivo REM con validación completa"""
    
    # Validar extensión
    if not file.filename.endswith('.xlsm'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsm")
    
    # Determinar programa_id
    if current_user["rol"] == "encargado":
        programa_id = current_user["programa_id"]
        if not programa_id:
            raise HTTPException(status_code=400, detail="Usuario sin programa asignado")
    
    if not programa_id:
        raise HTTPException(status_code=400, detail="programa_id requerido")
    
    # Validar formato de periodo
    if not re.match(r'^\d{4}-\d{2}$', periodo):
        raise HTTPException(status_code=400, detail="Formato de periodo inválido. Use YYYY-MM")
    
    # Verificar si ya existe archivo activo no-rechazado
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id FROM archivos 
        WHERE programa_id = ? AND periodo = ? AND activo = 1 AND estado != 'rechazado'
    """, programa_id, periodo)
    
    if cursor.fetchone():
        conn.close()
        raise HTTPException(
            status_code=400, 
            detail="Ya existe un archivo activo para este programa y período. Espere a que sea procesado o rechazado."
        )
    
    # Guardar temporalmente
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_filename = f"temp_{timestamp}_{file.filename}"
    temp_filepath = TEMP_UPLOAD_DIR / temp_filename
    
    with open(temp_filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Validar archivo
    validacion = validar_archivo_rem(temp_filepath, periodo)
    
    if not validacion["valido"]:
        temp_filepath.unlink()
        conn.close()
        raise HTTPException(status_code=400, detail=validacion)
    
    # Desactivar archivos anteriores
    cursor.execute("""
        UPDATE archivos 
        SET activo = 0 
        WHERE programa_id = ? AND periodo = ? AND activo = 1
    """, programa_id, periodo)
    
    # Mover a carpeta definitiva
    filename = f"{programa_id}_{timestamp}_{file.filename}"
    upload_dir_periodo = get_upload_dir_for_periodo(periodo)
    filepath = upload_dir_periodo / filename
    
    shutil.move(str(temp_filepath), str(filepath))
    
    # Insertar registro
    cursor.execute("""
        INSERT INTO archivos (usuario_id, programa_id, nombre_archivo, ruta_archivo, estado, periodo)
        VALUES (?, ?, ?, ?, 'pendiente', ?)
    """, current_user["id"], programa_id, file.filename, str(filepath), periodo)
    
    conn.commit()
    archivo_id = cursor.execute("SELECT @@IDENTITY").fetchone()[0]
    conn.close()
    
    registrar_log(
        current_user["id"], 
        "subida",
        f"Archivo: {file.filename} (Periodo: {periodo})", 
        archivo_id
    )
    
    return {
        "message": "Archivo subido exitosamente",
        "archivo_id": archivo_id,
        "filename": filename,
        "periodo": periodo
    }

@app.post("/archivos/validar")
def validar_archivo(
    request: ArchivoValidar,
    current_user: dict = Depends(get_current_user)
):
    """Validar o rechazar archivo (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if request.estado not in ["validado", "rechazado"]:
        raise HTTPException(status_code=400, detail="Estado debe ser 'validado' o 'rechazado'")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE archivos
        SET estado = ?,
            observaciones = ?,
            fecha_validacion = GETDATE(),
            validado_por = ?
        WHERE id = ?
    """, request.estado, request.observaciones, current_user["id"], request.archivo_id)
    
    conn.commit()
    conn.close()
    
    accion = "validacion" if request.estado == "validado" else "rechazo"
    
    registrar_log(
        current_user["id"],
        accion,
        request.observaciones or f"Archivo {request.estado}",
        request.archivo_id
    )
    
    return {"message": f"Archivo {request.estado} exitosamente"}

@app.get("/archivos/{archivo_id}/download")
def descargar_archivo(
    archivo_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Descargar archivo"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, ruta_archivo, nombre_archivo, usuario_id 
        FROM archivos 
        WHERE id = ?
    """, archivo_id)
    
    archivo = cursor.fetchone()
    conn.close()
    
    if not archivo:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    if current_user["rol"] == "encargado" and archivo[3] != current_user["id"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    filepath = Path(archivo[1])
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo físico no encontrado")
    
    return FileResponse(
        path=str(filepath),
        filename=archivo[2],
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

@app.post("/archivos/{archivo_id}/resubir")
async def resubir_archivo(
    archivo_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Re-subir archivo reparado (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if not file.filename.endswith('.xlsm'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsm")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT usuario_id, programa_id, ruta_archivo, periodo 
        FROM archivos 
        WHERE id = ?
    """, archivo_id)
    
    archivo = cursor.fetchone()
    
    if not archivo:
        conn.close()
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Eliminar archivo anterior
    old_filepath = Path(archivo[2])
    if old_filepath.exists():
        old_filepath.unlink()
    
    # Guardar nuevo archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{archivo[1]}_{timestamp}_{file.filename}"
    upload_dir_periodo = get_upload_dir_for_periodo(archivo[3])
    filepath = upload_dir_periodo / filename
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Actualizar registro
    cursor.execute("""
        UPDATE archivos
        SET nombre_archivo = ?,
            ruta_archivo = ?,
            estado = 'pendiente',
            observaciones = 'Documento reparado por administrador',
            fecha_validacion = NULL,
            validado_por = NULL
        WHERE id = ?
    """, file.filename, str(filepath), archivo_id)
    
    conn.commit()
    conn.close()
    
    registrar_log(
        current_user["id"],
        "resubida",
        f"Documento reparado: {file.filename}",
        archivo_id
    )
    
    return {
        "message": "Archivo re-subido exitosamente",
        "archivo_id": archivo_id,
        "filename": filename
    }

@app.get("/archivos/{archivo_id}/historial")
def historial_archivo(archivo_id: int, current_user: dict = Depends(get_current_user)):
    """Historial de actividad de un archivo específico"""
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar que el archivo existe
    cursor.execute("SELECT id, usuario_id FROM archivos WHERE id = ?", archivo_id)
    archivo = cursor.fetchone()
    
    if not archivo:
        conn.close()
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Si es encargado, solo puede ver historial de sus archivos
    if current_user["rol"] == "encargado" and archivo[1] != current_user["id"]:
        conn.close()
        raise HTTPException(status_code=403, detail="No tiene permiso para ver este historial")
    
    # Obtener historial
    cursor.execute("""
        SELECT l.id, l.accion, l.detalle, l.fecha, u.nombre as usuario_nombre
        FROM log_actividad l
        JOIN usuarios u ON l.usuario_id = u.id
        WHERE l.archivo_id = ?
        ORDER BY l.fecha ASC
    """, archivo_id)
    
    logs = []
    for row in cursor.fetchall():
        logs.append({
            "id": row[0],
            "accion": row[1],
            "detalle": row[2],
            "fecha": row[3],
            "usuario_nombre": row[4]
        })
    
    conn.close()
    return logs

# ==================== CONSOLIDACIONES ====================

@app.get("/consolidaciones", response_model=List[ConsolidacionInfo])
def listar_consolidaciones(
    periodo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar consolidaciones"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT 
            c.id, c.nombre_archivo, c.archivos_count,
            c.creado_por, c.fecha, c.periodo, u.nombre as creado_por_nombre
        FROM consolidaciones c
        JOIN usuarios u ON c.creado_por = u.id
    """
    params = []
    
    if periodo:
        query += " WHERE c.periodo = ?"
        params.append(periodo)
    
    query += " ORDER BY c.fecha DESC"
    
    if params:
        cursor.execute(query, *params)
    else:
        cursor.execute(query)
    
    consolidaciones = []
    for row in cursor.fetchall():
        consolidaciones.append({
            "id": row[0],
            "nombre_archivo": row[1],
            "archivos_count": row[2],
            "creado_por": row[3],
            "fecha": row[4],
            "periodo": row[5],
            "creado_por_nombre": row[6]
        })
    
    conn.close()
    return consolidaciones

@app.post("/consolidar")
def consolidar_archivos(
    request: ConsolidarRequest,
    current_user: dict = Depends(get_current_user)
):
    """Consolidar archivos validados"""
    # Validar permisos
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Validar cantidad mínima
    if len(request.archivos_ids) < 2:
        raise HTTPException(status_code=400, detail="Se requieren al menos 2 archivos")
    
    # CONVERTIR TODO A STRING EXPLÍCITAMENTE
    periodo_str = str(request.periodo) if request.periodo else ""
    
    # Validar formato de periodo
    if not re.match(r'^\d{4}-\d{2}$', periodo_str):
        raise HTTPException(status_code=400, detail="Formato de periodo inválido")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Verificar archivos
        placeholders = ','.join(['?' for _ in request.archivos_ids])
        query = f"""
            SELECT id, ruta_archivo, estado, periodo
            FROM archivos
            WHERE id IN ({placeholders}) AND activo = 1
        """
        cursor.execute(query, *request.archivos_ids)
        
        archivos = cursor.fetchall()
        
        if len(archivos) != len(request.archivos_ids):
            conn.close()
            raise HTTPException(status_code=400, detail="Algunos archivos no existen")
        
        # Validar estado y periodo
        for archivo in archivos:
            archivo_id = int(archivo[0])
            archivo_ruta = str(archivo[1])
            archivo_estado = str(archivo[2])
            archivo_periodo = str(archivo[3]) if archivo[3] else ""
            
            if archivo_periodo != periodo_str:
                conn.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"El archivo {archivo_id} es del periodo {archivo_periodo}, se esperaba {periodo_str}"
                )
            if archivo_estado != 'validado':
                conn.close()
                raise HTTPException(
                    status_code=400, 
                    detail=f"El archivo {archivo_id} no está validado (estado: {archivo_estado})"
                )
        
        # Proceso de consolidación
        plantilla_path = str(archivos[0][1])
        wb_consolidado = load_workbook(plantilla_path, keep_vba=True)
        
        hojas_datos = [sheet for sheet in wb_consolidado.sheetnames 
                      if sheet not in ['NOMBRE', 'Control', 'MACROS']]
        
        # Inicializar en 0
        for sheet_name in hojas_datos:
            ws = wb_consolidado[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.data_type == 'n' and not cell.protection.locked:
                        cell.value = 0
        
        # Sumar valores
        for archivo in archivos:
            archivo_path = str(archivo[1])
            wb_temp = load_workbook(archivo_path, data_only=True)
            
            for sheet_name in hojas_datos:
                if sheet_name not in wb_temp.sheetnames:
                    continue
                
                ws_temp = wb_temp[sheet_name]
                ws_consolidado = wb_consolidado[sheet_name]
                
                for row_idx, row in enumerate(ws_temp.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if isinstance(cell, MergedCell):
                            continue
                        
                        cell_consolidado = ws_consolidado.cell(row_idx, col_idx)
                        
                        if (cell.data_type == 'n' and 
                            not cell_consolidado.protection.locked and 
                            cell.value is not None):
                            
                            current_value = cell_consolidado.value or 0
                            cell_consolidado.value = current_value + (cell.value or 0)
            
            wb_temp.close()
        
        # Actualizar hoja NOMBRE
        if 'NOMBRE' in wb_consolidado.sheetnames:
            ws_nombre = wb_consolidado['NOMBRE']
            ws_nombre['B4'] = f"ARCHIVO CONSOLIDADO - Periodo {periodo_str}"
            ws_nombre['B5'] = f"Fecha consolidación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Guardar archivo
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"REM_Consolidado_{periodo_str}_{timestamp_str}.xlsm"
        upload_dir_periodo = get_upload_dir_for_periodo(periodo_str)
        filepath = upload_dir_periodo / filename
        filepath_str = str(filepath)
        
        wb_consolidado.save(filepath_str)
        wb_consolidado.close()
        
        # Registrar consolidación en BD
        archivos_count = len(request.archivos_ids)
        usuario_id = int(current_user["id"])
        
        cursor.execute("""
            INSERT INTO consolidaciones (nombre_archivo, ruta_archivo, archivos_count, creado_por, periodo)
            VALUES (?, ?, ?, ?, ?)
        """, filename, filepath_str, archivos_count, usuario_id, periodo_str)
        
        consolidacion_id = int(cursor.execute("SELECT @@IDENTITY").fetchone()[0])
        
        # Insertar relaciones y actualizar estados
        for archivo_id in request.archivos_ids:
            archivo_id_int = int(archivo_id)
            
            cursor.execute("""
                INSERT INTO consolidacion_archivos (consolidacion_id, archivo_id)
                VALUES (?, ?)
            """, consolidacion_id, archivo_id_int)
            
            cursor.execute("""
                UPDATE archivos 
                SET estado = 'consolidado' 
                WHERE id = ?
            """, archivo_id_int)
        
        conn.commit()
        conn.close()
        
        # Registrar logs
        detalle_general = f"{archivos_count} archivos (Periodo: {periodo_str})"
        registrar_log(
            usuario_id,
            "consolidacion",
            detalle_general,
            consolidacion_id=consolidacion_id
        )
        
        # Log por cada archivo
        for archivo_id in request.archivos_ids:
            archivo_id_int = int(archivo_id)
            detalle_archivo = f"Incluido en consolidación {consolidacion_id}"
            registrar_log(
                usuario_id,
                "consolidacion",
                detalle_archivo,
                archivo_id=archivo_id_int
            )
        
        return {
            "message": "Consolidación exitosa",
            "consolidacion_id": consolidacion_id,
            "archivo": filename,
            "periodo": periodo_str
        }
        
    except HTTPException:
        # Re-lanzar HTTPException sin modificar
        raise
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        print("="*70)
        print("❌ ERROR EN CONSOLIDACIÓN:")
        print(error_completo)
        print("="*70)
        conn.close()
        raise HTTPException(
            status_code=500, 
            detail=f"Error en consolidación: {str(e)}"
        )



@app.get("/consolidaciones/{consolidacion_id}/download")
def descargar_consolidacion(
    consolidacion_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Descargar archivo consolidado"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT nombre_archivo, ruta_archivo
        FROM consolidaciones
        WHERE id = ?
    """, consolidacion_id)
    
    consolidacion = cursor.fetchone()
    conn.close()
    
    if not consolidacion:
        raise HTTPException(status_code=404, detail="Consolidación no encontrada")
    
    filepath = Path(consolidacion[1])
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo físico no encontrado")
    
    return FileResponse(
        path=str(filepath),
        filename=consolidacion[0],
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

# ==================== LOGS ====================

@app.get("/logs")
def obtener_logs(
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Obtener logs de actividad (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(f"""
        SELECT TOP {limit}
            l.id, l.usuario_id, l.accion, l.detalle, l.fecha,
            u.nombre as usuario_nombre, l.archivo_id, l.consolidacion_id
        FROM log_actividad l
        JOIN usuarios u ON l.usuario_id = u.id
        ORDER BY l.fecha DESC
    """)
    
    logs = []
    for row in cursor.fetchall():
        logs.append({
            "id": row[0],
            "usuario_id": row[1],
            "accion": row[2],
            "detalle": row[3],
            "fecha": row[4],
            "usuario_nombre": row[5],
            "archivo_id": row[6],
            "consolidacion_id": row[7]
        })
    
    conn.close()
    return logs

# ==================== FIN ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
