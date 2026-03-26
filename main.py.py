"""
API REST para Sistema de Consolidación REM
FastAPI + SQL Server + pyodbc
Versión: 2.5.0
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, status
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime, timedelta
import pyodbc
import bcrypt
import jwt
import os
from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import re

# Cargar variables de entorno
from dotenv import load_dotenv
load_dotenv()

# ==================== CONFIGURACIÓN ====================
app = FastAPI(
    title="API Consolidador REM",
    description="API para gestión de archivos REM y consolidación con períodos mensuales",
    version="2.5.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración JWT
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise ValueError("❌ ERROR: SECRET_KEY no configurada en .env")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "480"))

# Configuración Base de Datos
DB_CONFIG = {
    "DRIVER": os.getenv("DB_DRIVER", "ODBC Driver 18 for SQL Server"),
    "SERVER": os.getenv("DB_SERVER", "172.25.5.70"),
    "DATABASE": os.getenv("DB_NAME", "Bi_Test"),
    "UID": os.getenv("DB_USER", "Bi_Estadistica"),
    "PWD": os.getenv("DB_PASSWORD", "EstadisticaBi2025.7"),
    "TrustServerCertificate": "yes",
    "Encrypt": "no"
}

# Directorio de archivos
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Directorio temporal para validación
TEMP_UPLOAD_DIR = Path("temp_uploads")
TEMP_UPLOAD_DIR.mkdir(exist_ok=True)

# Security
security = HTTPBearer()

print("============================================================")
print("🚀 API CONSOLIDADOR REM v2.5.0")
print("============================================================")
print("📊 CONFIGURACIÓN DE BASE DE DATOS:")
print(f"   SERVER: {DB_CONFIG['SERVER']}")
print(f"   DATABASE: {DB_CONFIG['DATABASE']}")
print(f"   DRIVER: {DB_CONFIG['DRIVER']}")
print("============================================================")

# ==================== MODELOS PYDANTIC ====================

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserRegister(BaseModel):
    nombre: str
    email: EmailStr
    password: str
    rol: str
    programa_id: Optional[int] = None

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

class UserInfo(BaseModel):
    id: int
    nombre: str
    email: str
    rol: str
    programa_id: Optional[int]
    activo: bool
    created_at: datetime
    programa_nombre: Optional[str]
    
    class Config:
        from_attributes = True

class ArchivoUpload(BaseModel):
    programa_id: int

class ArchivoInfo(BaseModel):
    id: int
    usuario_id: int
    programa_id: int
    nombre_archivo: str
    estado: str
    observaciones: Optional[str]
    fecha_subida: datetime
    fecha_validacion: Optional[datetime]
    validado_por: Optional[int]
    activo: bool
    usuario_nombre: str
    programa_nombre: str
    validado_por_nombre: Optional[str]
    periodo: str
    
    class Config:
        from_attributes = True

class ArchivoValidar(BaseModel):
    archivo_id: int
    estado: str
    observaciones: Optional[str] = None

class ConsolidacionCreate(BaseModel):
    archivos_ids: List[int]
    periodo: str

class ConsolidacionInfo(BaseModel):
    id: int
    nombre_archivo: str
    ruta_archivo: str
    fecha: datetime
    archivos_count: int
    creado_por: int
    creado_por_nombre: str
    periodo: str
    
    class Config:
        from_attributes = True

class ProgramaInfo(BaseModel):
    id: int
    nombre: str
    codigo: Optional[str]
    
    class Config:
        from_attributes = True

# ==================== FUNCIONES HELPER ====================

def get_db_connection():
    """Establecer conexión con SQL Server"""
    try:
        conn_str = (
            f"DRIVER={{{DB_CONFIG['DRIVER']}}};"
            f"SERVER={DB_CONFIG['SERVER']};"
            f"DATABASE={DB_CONFIG['DATABASE']};"
            f"UID={DB_CONFIG['UID']};"
            f"PWD={DB_CONFIG['PWD']};"
            f"TrustServerCertificate={DB_CONFIG['TrustServerCertificate']};"
            f"Encrypt={DB_CONFIG['Encrypt']};"
        )
        conn = pyodbc.connect(conn_str)
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error de conexión a BD: {str(e)}")

def registrar_log(usuario_id: int, accion: str, detalle: str = None, 
                  archivo_id: int = None, consolidacion_id: int = None):
    """
    Registrar actividad en log
    
    Args:
        usuario_id: ID del usuario que realiza la acción
        accion: Tipo de acción (subir, validar, rechazar, consolidar, etc.)
        detalle: Descripción detallada de la acción
        archivo_id: ID del archivo relacionado (opcional)
        consolidacion_id: ID de la consolidación relacionada (opcional)
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO log_actividad (usuario_id, accion, detalle, archivo_id, consolidacion_id)
            VALUES (?, ?, ?, ?, ?)
        """, usuario_id, accion, detalle, archivo_id, consolidacion_id)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠️  Error registrando log: {e}")

def get_periodo_actual() -> str:
    """Retorna el período actual en formato YYYY-MM"""
    now = datetime.now()
    return now.strftime("%Y-%m")

def get_upload_dir_for_periodo(periodo: str) -> Path:
    """
    Retorna el directorio de uploads para un período específico
    Crea el directorio si no existe
    """
    periodo_dir = UPLOAD_DIR / periodo
    periodo_dir.mkdir(exist_ok=True)
    return periodo_dir

def validar_mes_archivo(filepath: Path) -> dict:
    """
    Valida que el archivo Excel corresponda al mes anterior
    
    Args:
        filepath: Ruta al archivo .xlsm
        
    Returns:
        dict con resultado de validación
    """
    # Mapeo de meses español
    MESES_MAP = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    
    try:
        # Obtener fecha actual
        ahora = datetime.now()
        mes_actual = ahora.month
        anio_actual = ahora.year
        
        # Calcular mes y año anterior
        if mes_actual == 1:
            mes_esperado = 12
            anio_esperado = anio_actual - 1
        else:
            mes_esperado = mes_actual - 1
            anio_esperado = anio_actual
        
        mes_esperado_nombre = MESES_MAP[mes_esperado]
        
        # Leer archivo Excel
        wb = load_workbook(filepath, data_only=True)
        
        if 'NOMBRE' not in wb.sheetnames:
            return {
                "valido": False,
                "mensaje": "El archivo no contiene la hoja NOMBRE requerida",
                "mes_archivo": None,
                "anio_archivo": None,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        ws = wb['NOMBRE']
        
        # Leer mes y año del archivo
        mes_archivo_valor = ws['B6'].value
        anio_archivo_valor = ws['B7'].value
        
        wb.close()
        
        # Validar que los campos tengan valores
        if not mes_archivo_valor:
            return {
                "valido": False,
                "mensaje": f"El campo B6 (Mes) está vacío. Debe contener: {mes_esperado_nombre}",
                "mes_archivo": None,
                "anio_archivo": anio_archivo_valor,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        if not anio_archivo_valor:
            return {
                "valido": False,
                "mensaje": f"El campo B7 (Año) está vacío. Debe contener: {anio_esperado}",
                "mes_archivo": mes_archivo_valor,
                "anio_archivo": None,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        # Normalizar mes del archivo
        mes_archivo_str = str(mes_archivo_valor).strip().upper()
        
        # Convertir año a entero
        try:
            anio_archivo_int = int(anio_archivo_valor)
        except (ValueError, TypeError):
            return {
                "valido": False,
                "mensaje": f"El año '{anio_archivo_valor}' en B7 no es válido",
                "mes_archivo": mes_archivo_str,
                "anio_archivo": None,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        # Validar mes
        if mes_archivo_str != mes_esperado_nombre:
            return {
                "valido": False,
                "mensaje": f"El mes del archivo es '{mes_archivo_str}' pero se esperaba '{mes_esperado_nombre}' (mes anterior al actual)",
                "mes_archivo": mes_archivo_str,
                "anio_archivo": anio_archivo_int,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        # Validar año
        if anio_archivo_int != anio_esperado:
            return {
                "valido": False,
                "mensaje": f"El año del archivo es {anio_archivo_int} pero se esperaba {anio_esperado}",
                "mes_archivo": mes_archivo_str,
                "anio_archivo": anio_archivo_int,
                "mes_esperado": mes_esperado_nombre,
                "anio_esperado": anio_esperado
            }
        
        # Todo OK
        return {
            "valido": True,
            "mensaje": f"Archivo válido: {mes_archivo_str} {anio_archivo_int}",
            "mes_archivo": mes_archivo_str,
            "anio_archivo": anio_archivo_int,
            "mes_esperado": mes_esperado_nombre,
            "anio_esperado": anio_esperado
        }
        
    except Exception as e:
        return {
            "valido": False,
            "mensaje": f"Error al validar archivo: {str(e)}",
            "mes_archivo": None,
            "anio_archivo": None,
            "mes_esperado": None,
            "anio_esperado": None
        }

# ==================== SEGURIDAD JWT ====================

def create_access_token(data: dict):
    """Crear token JWT"""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Obtener usuario actual desde token JWT"""
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token inválido")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT u.id, u.nombre, u.email, u.rol, u.programa_id, p.nombre as programa_nombre
            FROM usuarios u
            LEFT JOIN programas p ON u.programa_id = p.id
            WHERE u.id = ? AND u.activo = 1
        """, user_id)
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        
        return {
            "id": user[0],
            "nombre": user[1],
            "email": user[2],
            "rol": user[3],
            "programa_id": user[4],
            "programa_nombre": user[5]
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ==================== ENDPOINTS ====================

# ==================== HEALTH CHECK ====================

@app.get("/health")
def health_check():
    """Health check endpoint"""
    try:
        conn = get_db_connection()
        conn.close()
        return {
            "status": "healthy",
            "version": "2.5.0",
            "database": "connected",
            "periodo_actual": get_periodo_actual()
        }
    except:
        return {
            "status": "unhealthy",
            "version": "2.5.0",
            "database": "disconnected"
        }

# ==================== AUTENTICACIÓN ====================

@app.post("/auth/login", response_model=Token)
def login(user: UserLogin):
    """Login de usuario"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.password, u.rol, u.programa_id, u.activo, p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        WHERE u.email = ?
    """, user.email)
    
    db_user = cursor.fetchone()
    conn.close()
    
    if not db_user:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    if not db_user[6]:
        raise HTTPException(status_code=401, detail="Usuario inactivo")
    
    if not bcrypt.checkpw(user.password.encode('utf-8'), db_user[3].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    access_token = create_access_token({"user_id": db_user[0]})
    
    registrar_log(db_user[0], "login", f"Inicio de sesión exitoso")
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": db_user[0],
            "nombre": db_user[1],
            "email": db_user[2],
            "rol": db_user[4],
            "programa_id": db_user[5],
            "programa_nombre": db_user[7]
        }
    }

@app.post("/auth/register")
def register(user: UserRegister, current_user: dict = Depends(get_current_user)):
    """Registrar nuevo usuario (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id FROM usuarios WHERE email = ?", user.email)
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Email ya registrado")
    
    hashed_password = bcrypt.hashpw(user.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    cursor.execute("""
        INSERT INTO usuarios (nombre, email, password, rol, programa_id)
        VALUES (?, ?, ?, ?, ?)
    """, user.nombre, user.email, hashed_password, user.rol, user.programa_id)
    
    conn.commit()
    nuevo_id = cursor.execute("SELECT @@IDENTITY").fetchone()[0]
    conn.close()
    
    registrar_log(current_user["id"], "register", f"Usuario creado: {user.email}")
    
    return {"message": "Usuario creado exitosamente", "user_id": nuevo_id}

@app.get("/auth/me", response_model=UserInfo)
def get_me(current_user: dict = Depends(get_current_user)):
    """Obtener información del usuario actual"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.rol, u.programa_id, u.activo, u.created_at, p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        WHERE u.id = ?
    """, current_user["id"])
    
    user = cursor.fetchone()
    conn.close()
    
    return {
        "id": user[0],
        "nombre": user[1],
        "email": user[2],
        "rol": user[3],
        "programa_id": user[4],
        "activo": user[5],
        "created_at": user[6],
        "programa_nombre": user[7]
    }

@app.get("/usuarios", response_model=List[UserInfo])
def listar_usuarios(current_user: dict = Depends(get_current_user)):
    """Listar usuarios (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.id, u.nombre, u.email, u.rol, u.programa_id, u.activo, u.created_at, p.nombre as programa_nombre
        FROM usuarios u
        LEFT JOIN programas p ON u.programa_id = p.id
        ORDER BY u.created_at DESC
    """)
    
    usuarios = []
    for row in cursor.fetchall():
        usuarios.append({
            "id": row[0],
            "nombre": row[1],
            "email": row[2],
            "rol": row[3],
            "programa_id": row[4],
            "activo": row[5],
            "created_at": row[6],
            "programa_nombre": row[7]
        })
    
    conn.close()
    return usuarios

# ==================== PROGRAMAS ====================

@app.get("/programas", response_model=List[ProgramaInfo])
def listar_programas():
    """Listar programas disponibles"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, nombre, codigo FROM programas ORDER BY nombre")
    
    programas = []
    for row in cursor.fetchall():
        programas.append({
            "id": row[0],
            "nombre": row[1],
            "codigo": row[2]
        })
    
    conn.close()
    return programas

# ==================== PERÍODOS ====================

@app.get("/periodo-actual")
def obtener_periodo_actual():
    """Obtener información del período actual"""
    periodo = get_periodo_actual()
    año, mes = periodo.split('-')
    
    meses_nombres = {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
    }
    
    return {
        "periodo": periodo,
        "mes": meses_nombres.get(mes, mes),
        "anio": int(año),
        "mes_numero": int(mes)
    }

@app.get("/periodos")
def obtener_periodos(current_user: dict = Depends(get_current_user)):
    """Obtener estadísticas de todos los períodos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Obtener periodos únicos
    cursor.execute("""
        SELECT DISTINCT periodo 
        FROM archivos 
        WHERE periodo IS NOT NULL
        ORDER BY periodo DESC
    """)
    
    periodos_list = []
    for row in cursor.fetchall():
        periodo = row[0]
        
        # Estadísticas del período
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) as pendientes,
                SUM(CASE WHEN estado = 'validado' THEN 1 ELSE 0 END) as validados,
                SUM(CASE WHEN estado = 'rechazado' THEN 1 ELSE 0 END) as rechazados,
                SUM(CASE WHEN estado = 'consolidado' THEN 1 ELSE 0 END) as consolidados
            FROM archivos
            WHERE periodo = ? AND activo = 1
        """, periodo)
        
        stats = cursor.fetchone()
        
        # Verificar si existe consolidación
        cursor.execute("""
            SELECT COUNT(*) FROM consolidaciones WHERE periodo = ?
        """, periodo)
        
        hay_consolidacion = cursor.fetchone()[0] > 0
        
        periodos_list.append({
            "periodo": periodo,
            "total_archivos": stats[0] or 0,
            "pendientes": stats[1] or 0,
            "validados": stats[2] or 0,
            "rechazados": stats[3] or 0,
            "consolidados": stats[4] or 0,
            "hay_consolidacion": hay_consolidacion,
            "puede_consolidar": (stats[2] or 0) > 0 and not hay_consolidacion
        })
    
    conn.close()
    return periodos_list

# ==================== PLANTILLA ====================

@app.get("/plantilla/download")
def descargar_plantilla():
    """Descargar plantilla base SA_26_V1.2.xlsm (público, sin autenticación)"""
    plantilla_path = Path("SA_26_V1.2.xlsm")
    
    if not plantilla_path.exists():
        raise HTTPException(
            status_code=404, 
            detail="Plantilla no encontrada. Coloque SA_26_V1.2.xlsm en la raíz del proyecto"
        )
    
    return FileResponse(
        path=str(plantilla_path),
        filename="SA_26_V1.2.xlsm",
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )


# ==================== ARCHIVOS ====================

@app.post("/archivos/upload")
async def subir_archivo(
    file: UploadFile = File(...),
    programa_id: int = None,
    periodo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Subir archivo REM con validación de mes anterior
    
    Args:
        file: Archivo .xlsm
        programa_id: ID del programa (encargados no lo necesitan)
        periodo: Periodo YYYY-MM (opcional, por defecto mes actual)
    """
    
    if not file.filename.endswith('.xlsm'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsm")
    
    if current_user["rol"] == "encargado":
        programa_id = current_user["programa_id"]
        if not programa_id:
            raise HTTPException(status_code=400, detail="Usuario sin programa asignado")
    
    if not programa_id:
        raise HTTPException(status_code=400, detail="programa_id requerido")
    
    # Obtener periodo (usar el enviado o el actual)
    periodo_archivo = periodo if periodo else get_periodo_actual()
    
    # Validar formato de periodo si fue enviado
    if periodo:
        if not re.match(r'^\d{4}-\d{2}$', periodo):
            raise HTTPException(
                status_code=400, 
                detail="Formato de periodo inválido. Use YYYY-MM (ej: 2026-03)"
            )
    
    # ============================================================
    # VALIDACIÓN DE MES ANTERIOR (solo si NO se envió periodo)
    # ============================================================
    
    # Guardar archivo temporalmente para validación
    temp_dir = Path("temp_uploads")
    temp_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_filename = f"temp_{timestamp}_{file.filename}"
    temp_filepath = temp_dir / temp_filename
    
    # Guardar temporalmente
    with open(temp_filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Si el frontend envió periodo explícito, NO validar mes
    # Esto permite subir archivos atrasados o de periodos específicos
    validar_mes = not periodo  # Validar solo si NO se envió periodo
    
    if validar_mes:
        # Validar que corresponda al mes anterior
        validacion = validar_mes_archivo(temp_filepath)
        
        if not validacion["valido"]:
            # Eliminar archivo temporal
            temp_filepath.unlink()
            
            # Retornar error descriptivo
            raise HTTPException(
                status_code=400, 
                detail={
                    "error": "Validación de mes fallida",
                    "mensaje": validacion["mensaje"],
                    "mes_archivo": validacion["mes_archivo"],
                    "anio_archivo": validacion["anio_archivo"],
                    "mes_esperado": validacion["mes_esperado"],
                    "anio_esperado": validacion["anio_esperado"],
                    "ayuda": f"El archivo debe corresponder al mes anterior. Si estamos en {datetime.now().strftime('%B %Y')}, el archivo debe ser de {validacion['mes_esperado']} {validacion['anio_esperado']}"
                }
            )
    else:
        # No validar mes si se envió periodo explícito
        # Solo leer la información del archivo para logging
        try:
            wb = load_workbook(temp_filepath, data_only=True)
            ws = wb['NOMBRE']
            mes_archivo = str(ws['B6'].value).strip().upper() if ws['B6'].value else "N/A"
            anio_archivo = int(ws['B7'].value) if ws['B7'].value else 0
            wb.close()
            
            validacion = {
                "valido": True,
                "mensaje": f"Periodo explícito: {periodo_archivo}. Archivo: {mes_archivo} {anio_archivo}",
                "mes_archivo": mes_archivo,
                "anio_archivo": anio_archivo
            }
        except:
            validacion = {
                "valido": True,
                "mensaje": f"Periodo explícito: {periodo_archivo}",
                "mes_archivo": "N/A",
                "anio_archivo": 0
            }
    
    # ============================================================
    # ARCHIVO VÁLIDO - CONTINUAR CON PROCESO NORMAL
    # ============================================================
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Desactivar archivo del mismo programa+periodo
    cursor.execute("""
        UPDATE archivos 
        SET activo = 0 
        WHERE programa_id = ? AND usuario_id = ? AND periodo = ? AND activo = 1
    """, programa_id, current_user["id"], periodo_archivo)
    
    # Mover de temporal a carpeta definitiva
    filename = f"{programa_id}_{timestamp}_{file.filename}"
    upload_dir_periodo = get_upload_dir_for_periodo(periodo_archivo)
    filepath = upload_dir_periodo / filename
    
    # Mover archivo
    shutil.move(str(temp_filepath), str(filepath))
    
    # Incluir periodo en INSERT
    cursor.execute("""
        INSERT INTO archivos (usuario_id, programa_id, nombre_archivo, ruta_archivo, estado, periodo)
        VALUES (?, ?, ?, ?, 'pendiente', ?)
    """, current_user["id"], programa_id, file.filename, str(filepath), periodo_archivo)
    
    conn.commit()
    archivo_id = cursor.execute("SELECT @@IDENTITY").fetchone()[0]
    conn.close()
    
    registrar_log(
        current_user["id"], 
        "subir", 
        f"Archivo: {file.filename} (Periodo: {periodo_archivo}, Mes validado: {validacion['mes_archivo']} {validacion['anio_archivo']})", 
        archivo_id
    )
    
    return {
        "message": "Archivo subido exitosamente",
        "archivo_id": archivo_id,
        "filename": filename,
        "periodo": periodo_archivo,
        "validacion": {
            "mes_archivo": validacion["mes_archivo"],
            "anio_archivo": validacion["anio_archivo"],
            "mensaje": validacion["mensaje"]
        }
    }

@app.get("/archivos", response_model=List[ArchivoInfo])
def listar_archivos(
    estado: Optional[str] = None,
    programa_id: Optional[int] = None,
    periodo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar archivos con filtros opcionales"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Si no se especifica periodo, usar el actual
    if periodo is None:
        periodo = get_periodo_actual()
    
    query = """
        SELECT 
            a.id, a.usuario_id, a.programa_id, a.nombre_archivo, a.estado,
            a.observaciones, a.fecha_subida, a.fecha_validacion, a.validado_por,
            a.activo, u.nombre as usuario_nombre, p.nombre as programa_nombre,
            v.nombre as validado_por_nombre, a.periodo
        FROM archivos a
        JOIN usuarios u ON a.usuario_id = u.id
        JOIN programas p ON a.programa_id = p.id
        LEFT JOIN usuarios v ON a.validado_por = v.id
        WHERE a.activo = 1
    """
    params = []
    
    # Filtro por periodo
    query += " AND a.periodo = ?"
    params.append(periodo)
    
    if estado:
        query += " AND a.estado = ?"
        params.append(estado)
    
    if programa_id:
        query += " AND a.programa_id = ?"
        params.append(programa_id)
    
    if current_user["rol"] == "encargado":
        query += " AND a.usuario_id = ?"
        params.append(current_user["id"])
    
    query += " ORDER BY a.fecha_subida DESC"
    
    cursor.execute(query, *params)
    
    archivos = []
    for row in cursor.fetchall():
        archivos.append({
            "id": row[0],
            "usuario_id": row[1],
            "programa_id": row[2],
            "nombre_archivo": row[3],
            "estado": row[4],
            "observaciones": row[5],
            "fecha_subida": row[6],
            "fecha_validacion": row[7],
            "validado_por": row[8],
            "activo": row[9],
            "usuario_nombre": row[10],
            "programa_nombre": row[11],
            "validado_por_nombre": row[12],
            "periodo": row[13]
        })
    
    conn.close()
    return archivos

@app.get("/archivos/{archivo_id}/download")
def descargar_archivo(
    archivo_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Descargar archivo"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, ruta_archivo, nombre_archivo, usuario_id 
        FROM archivos 
        WHERE id = ?
    """, archivo_id)
    
    archivo = cursor.fetchone()
    conn.close()
    
    if not archivo:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    if current_user["rol"] == "encargado" and archivo[3] != current_user["id"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    filepath = Path(archivo[1])
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo físico no encontrado")
    
    registrar_log(
        current_user["id"],
        "descargar",
        f"Archivo ID {archivo_id}: {archivo[2]}",
        archivo_id
    )
    
    return FileResponse(
        path=str(filepath),
        filename=archivo[2],
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

@app.get("/archivos/{archivo_id}/historial")
def obtener_historial_archivo(
    archivo_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Obtener historial de acciones sobre un archivo específico"""
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar que el archivo existe
    cursor.execute("SELECT id, usuario_id FROM archivos WHERE id = ?", archivo_id)
    archivo = cursor.fetchone()
    
    if not archivo:
        conn.close()
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Si es encargado, solo puede ver historial de sus propios archivos
    if current_user["rol"] == "encargado" and archivo[1] != current_user["id"]:
        conn.close()
        raise HTTPException(status_code=403, detail="No tiene permiso para ver este historial")
    
    # Obtener historial del archivo
    cursor.execute("""
        SELECT l.id, l.accion, l.detalle, l.fecha, u.nombre as usuario_nombre
        FROM log_actividad l
        JOIN usuarios u ON l.usuario_id = u.id
        WHERE l.archivo_id = ?
        ORDER BY l.fecha ASC
    """, archivo_id)
    
    historial = []
    for row in cursor.fetchall():
        historial.append({
            "id": row[0],
            "accion": row[1],
            "detalle": row[2],
            "fecha": row[3],
            "usuario_nombre": row[4]
        })
    
    conn.close()
    return historial

@app.post("/archivos/{archivo_id}/resubir")
async def resubir_archivo(
    archivo_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Re-subir archivo (admin para reparar archivos rechazados)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if not file.filename.endswith('.xlsm'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsm")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Obtener info del archivo original
    cursor.execute("""
        SELECT usuario_id, programa_id, ruta_archivo, periodo 
        FROM archivos 
        WHERE id = ?
    """, archivo_id)
    
    archivo_original = cursor.fetchone()
    
    if not archivo_original:
        conn.close()
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Eliminar archivo físico anterior
    old_filepath = Path(archivo_original[2])
    if old_filepath.exists():
        old_filepath.unlink()
    
    # Guardar nuevo archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{archivo_original[1]}_{timestamp}_{file.filename}"
    upload_dir_periodo = get_upload_dir_for_periodo(archivo_original[3])
    filepath = upload_dir_periodo / filename
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Actualizar registro
    cursor.execute("""
        UPDATE archivos
        SET nombre_archivo = ?,
            ruta_archivo = ?,
            estado = 'pendiente',
            observaciones = 'Documento reparado por administrador',
            fecha_validacion = NULL,
            validado_por = NULL
        WHERE id = ?
    """, file.filename, str(filepath), archivo_id)
    
    conn.commit()
    conn.close()
    
    registrar_log(
        current_user["id"],
        "resubir",
        f"Archivo ID {archivo_id}: Documento reparado por administrador",
        archivo_id
    )
    
    return {
        "message": "Archivo re-subido exitosamente",
        "archivo_id": archivo_id,
        "filename": filename
    }

@app.post("/archivos/validar")
def validar_archivo(
    request: ArchivoValidar,
    current_user: dict = Depends(get_current_user)
):
    """Validar o rechazar archivo (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if request.estado not in ["validado", "rechazado"]:
        raise HTTPException(status_code=400, detail="Estado debe ser 'validado' o 'rechazado'")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE archivos
        SET estado = ?,
            observaciones = ?,
            fecha_validacion = GETDATE(),
            validado_por = ?
        WHERE id = ?
    """, request.estado, request.observaciones, current_user["id"], request.archivo_id)
    
    conn.commit()
    conn.close()
    
    accion = "validar" if request.estado == "validado" else "rechazar"
    
    registrar_log(
        current_user["id"],
        accion,
        f"Archivo ID {request.archivo_id}: {request.observaciones or 'Sin observaciones'}",
        request.archivo_id
    )
    
    return {"message": f"Archivo {request.estado} exitosamente"}

@app.post("/archivos/validar-mes")
async def validar_mes_antes_de_subir(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Valida que el archivo corresponda al mes anterior SIN subirlo
    Útil para el frontend para validar antes de enviar
    """
    
    if not file.filename.endswith('.xlsm'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsm")
    
    # Guardar temporalmente
    temp_dir = Path("temp_uploads")
    temp_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_filename = f"temp_val_{timestamp}_{file.filename}"
    temp_filepath = temp_dir / temp_filename
    
    with open(temp_filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Validar
    validacion = validar_mes_archivo(temp_filepath)
    
    # Eliminar archivo temporal
    temp_filepath.unlink()
    
    # Retornar resultado
    return {
        "valido": validacion["valido"],
        "mensaje": validacion["mensaje"],
        "mes_archivo": validacion["mes_archivo"],
        "anio_archivo": validacion["anio_archivo"],
        "mes_esperado": validacion["mes_esperado"],
        "anio_esperado": validacion["anio_esperado"],
        "fecha_actual": datetime.now().strftime("%B %Y")
    }


# ==================== CONSOLIDACIÓN ====================

@app.post("/consolidar")
def consolidar_archivos(
    request: ConsolidacionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Consolidar archivos validados en un solo archivo"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    if len(request.archivos_ids) < 2:
        raise HTTPException(status_code=400, detail="Se requieren al menos 2 archivos")
    
    # Validar formato de período
    if not re.match(r'^\d{4}-\d{2}$', request.periodo):
        raise HTTPException(
            status_code=400,
            detail="Formato de periodo inválido. Use YYYY-MM (ej: 2026-03)"
        )
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar que todos los archivos existen, están validados y son del mismo periodo
    placeholders = ','.join('?' * len(request.archivos_ids))
    cursor.execute(f"""
        SELECT id, ruta_archivo, estado, periodo
        FROM archivos
        WHERE id IN ({placeholders}) AND activo = 1
    """, *request.archivos_ids)
    
    archivos = cursor.fetchall()
    
    if len(archivos) != len(request.archivos_ids):
        conn.close()
        raise HTTPException(status_code=400, detail="Algunos archivos no existen")
    
    # Verificar que todos son del periodo solicitado
    for archivo in archivos:
        if archivo[3] != request.periodo:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"El archivo {archivo[0]} es del periodo {archivo[3]}, se esperaba {request.periodo}"
            )
        if archivo[2] != 'validado':
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"El archivo {archivo[0]} no está validado"
            )
    
    # Consolidar archivos
    try:
        # Usar primer archivo como plantilla
        plantilla_path = archivos[0][1]
        wb_consolidado = load_workbook(plantilla_path, keep_vba=True)
        
        # Identificar hojas de datos
        hojas_datos = [sheet for sheet in wb_consolidado.sheetnames 
                      if sheet not in ['NOMBRE', 'Control', 'MACROS']]
        
        # Inicializar celdas editables en 0
        for sheet_name in hojas_datos:
            ws = wb_consolidado[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.data_type == 'n' and not cell.protection.locked:
                        cell.value = 0
        
        # Sumar valores de todos los archivos
        for archivo_path in [a[1] for a in archivos]:
            wb_temp = load_workbook(archivo_path, data_only=True)
            
            for sheet_name in hojas_datos:
                if sheet_name not in wb_temp.sheetnames:
                    continue
                
                ws_temp = wb_temp[sheet_name]
                ws_consolidado = wb_consolidado[sheet_name]
                
                for row_idx, row in enumerate(ws_temp.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if isinstance(cell, MergedCell):
                            continue
                        
                        cell_consolidado = ws_consolidado.cell(row_idx, col_idx)
                        
                        if (cell.data_type == 'n' and 
                            not cell_consolidado.protection.locked and 
                            cell.value is not None):
                            
                            current_value = cell_consolidado.value or 0
                            cell_consolidado.value = current_value + (cell.value or 0)
            
            wb_temp.close()
        
        # Actualizar hoja NOMBRE
        if 'NOMBRE' in wb_consolidado.sheetnames:
            ws_nombre = wb_consolidado['NOMBRE']
            ws_nombre['B4'] = f"ARCHIVO CONSOLIDADO - Periodo {request.periodo}"
            ws_nombre['B5'] = f"Fecha consolidación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Guardar archivo consolidado
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"REM_Consolidado_{request.periodo}_{timestamp}.xlsm"
        upload_dir_periodo = get_upload_dir_for_periodo(request.periodo)
        filepath = upload_dir_periodo / filename
        
        wb_consolidado.save(filepath)
        wb_consolidado.close()
        
        # Registrar en BD
        cursor.execute("""
            INSERT INTO consolidaciones (nombre_archivo, ruta_archivo, creado_por, periodo)
            VALUES (?, ?, ?, ?)
        """, filename, str(filepath), current_user["id"], request.periodo)
        
        consolidacion_id = cursor.execute("SELECT @@IDENTITY").fetchone()[0]
        
        # Registrar archivos consolidados
        for archivo_id in request.archivos_ids:
            cursor.execute("""
                INSERT INTO consolidacion_archivos (consolidacion_id, archivo_id)
                VALUES (?, ?)
            """, consolidacion_id, archivo_id)
            
            # Marcar archivo como consolidado
            cursor.execute("""
                UPDATE archivos
                SET estado = 'consolidado'
                WHERE id = ?
            """, archivo_id)
        
        conn.commit()
        conn.close()
        
        # Registrar log general
        registrar_log(
            current_user["id"],
            "consolidar",
            f"{len(request.archivos_ids)} archivos (Periodo: {request.periodo})",
            consolidacion_id=consolidacion_id
        )
        
        # Registrar log para cada archivo consolidado
        for archivo_id in request.archivos_ids:
            registrar_log(
                current_user["id"],
                "consolidar",
                f"Incluido en consolidación {consolidacion_id} del periodo {request.periodo}",
                archivo_id
            )
        
        return {
            "message": "Consolidación exitosa",
            "consolidacion_id": consolidacion_id,
            "archivo": filename,
            "periodo": request.periodo
        }
        
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Error en consolidación: {str(e)}")

@app.get("/consolidaciones", response_model=List[ConsolidacionInfo])
def listar_consolidaciones(
    periodo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar consolidaciones"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT 
            c.id, c.nombre_archivo, c.ruta_archivo, c.fecha,
            c.creado_por, u.nombre as creado_por_nombre, c.periodo,
            COUNT(ca.archivo_id) as archivos_count
        FROM consolidaciones c
        JOIN usuarios u ON c.creado_por = u.id
        LEFT JOIN consolidacion_archivos ca ON c.id = ca.consolidacion_id
    """
    params = []
    
    if periodo:
        query += " WHERE c.periodo = ?"
        params.append(periodo)
    
    query += " GROUP BY c.id, c.nombre_archivo, c.ruta_archivo, c.fecha, c.creado_por, u.nombre, c.periodo"
    query += " ORDER BY c.fecha DESC"
    
    cursor.execute(query, *params)
    
    consolidaciones = []
    for row in cursor.fetchall():
        consolidaciones.append({
            "id": row[0],
            "nombre_archivo": row[1],
            "ruta_archivo": row[2],
            "fecha": row[3],
            "creado_por": row[4],
            "creado_por_nombre": row[5],
            "periodo": row[6],
            "archivos_count": row[7]
        })
    
    conn.close()
    return consolidaciones

@app.get("/consolidaciones/{consolidacion_id}/download")
def descargar_consolidacion(
    consolidacion_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Descargar archivo consolidado"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT nombre_archivo, ruta_archivo
        FROM consolidaciones
        WHERE id = ?
    """, consolidacion_id)
    
    consolidacion = cursor.fetchone()
    conn.close()
    
    if not consolidacion:
        raise HTTPException(status_code=404, detail="Consolidación no encontrada")
    
    filepath = Path(consolidacion[1])
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo físico no encontrado")
    
    registrar_log(
        current_user["id"],
        "descargar",
        f"Consolidación ID {consolidacion_id}",
        consolidacion_id=consolidacion_id
    )
    
    return FileResponse(
        path=str(filepath),
        filename=consolidacion[0],
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

# ==================== LOGS ====================

@app.get("/logs")
def obtener_logs(
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Obtener logs de actividad (solo admin)"""
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(f"""
        SELECT TOP {limit}
            l.id, l.usuario_id, l.accion, l.detalle, l.fecha,
            u.nombre as usuario_nombre, l.archivo_id, l.consolidacion_id
        FROM log_actividad l
        JOIN usuarios u ON l.usuario_id = u.id
        ORDER BY l.fecha DESC
    """)
    
    logs = []
    for row in cursor.fetchall():
        logs.append({
            "id": row[0],
            "usuario_id": row[1],
            "accion": row[2],
            "detalle": row[3],
            "fecha": row[4],
            "usuario_nombre": row[5],
            "archivo_id": row[6],
            "consolidacion_id": row[7]
        })
    
    conn.close()
    return logs

# ==================== FIN ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
